from flask import Flask, request
from linebot.models import TemplateSendMessage,ButtonsTemplate,MessageAction,URIAction,TextSendMessage,ImageSendMessage,CarouselTemplate,CarouselColumn
import json
from ping3 import ping 
import os
from linebot import LineBotApi, WebhookHandler
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import keyboard 
import time
import requests
import re
import random
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from pylibdmtx.pylibdmtx import encode
from drive_upload import upload_to_drive
from flask import Flask, request, abort
from linebot.models import *
from linebot.v3 import WebhookParser
from linebot.v3.webhooks import MessageEvent, TextMessageContent
model_path = "/Users/chienchihhsiang/.lmstudio/models/shenzhi-wang/Llama3-8B-Chinese-Chat-GGUF-8bit/Llama3-8B-Chinese-Chat-q8_0-v2_1.gguf"
LLAMA_API_URL = 'http://192.168.1.109:1234/v1/chat/completions' 
Material = False
in_process = False
in_process2 = False
pass_process = False
sec_in_process = False
sec_in_process2 = False
unlock_repaire_station = False
image_process = False
unlock_repaire_station2 = False
QRcode_check = False
return_msg = False
internet_process = False
class_process = False
global error_code
error_code = None
flow1 = False
flow2 = False
work_flow1 = False
work_flow2 = False
line_bot_api = LineBotApi('Pn6oyBEK0RJLR1U0OEY7q35O+RTM0PaHrSAzQ8805ZRqNqf/O2A1CM5YWMeQoB+8ngYCLqoEK6nmHi9T6YCycuSOSKHiTBQWG56vUUY6RTbDr96z2Iq+Uo7SSVdWWBuDOdClGrB4Y8bfvXrMqmnwBQdB04t89/1O/w1cDnyilFU=')
handler = WebhookHandler('7599d376613784d49a0230e1b53a7eaf')
chrome_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
AT = "U1d2fae8c03fbc0819fe079da5250d9bc"

def check_or_create_excel(file_path):
    if os.path.exists(file_path):
        print(f"檔案 '{file_path}' 已存在，正在讀取...")
        workbook = load_workbook(file_path)
        sheet = workbook.active
        if sheet.max_row > 1 and sheet.max_column > 1:
            print("第一列和第一行的數據如下：")
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    print(cell.value, end="\t")
            print()
        else:
            print("檔案存在，但內容是空的。")
    else:
        print(f"檔案 '{file_path}' 不存在，正在創建新檔案...")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Log"
        sheet.append(["Timestamp", "ID", "Event", "Department", "Class", "Serial Number", "Result_output"])
        print("已創建標題行。")
        os.makedirs(os.path.dirname(file_path), exist_ok=True) 
        workbook.save(file_path)
        print(f"新檔案 '{file_path}' 已創建並儲存。")
    print("操作完成！")

def repair_ac2_automation(SN_input):
    global actions2
    global error_code
    global error_text
    error_text = None
    error_code = False
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument('--headless')    
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })     
    options.binary_location = chrome_path 
    input_acc = "14575"
    serice = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=serice)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        global last_column
        driver.get(f"http://cimes.seec.com.tw/AMG/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule"]')))
        element4.click()
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '維修(客)')]")))
        actions.move_to_element(button2).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        input1 = wait.until(EC.presence_of_element_located((By.ID, "ttbLot")))
        input1.send_keys(SN_input)
        input1.send_keys(Keys.RETURN)
        time.sleep(1)
        select_OTHER = wait.until(EC.presence_of_element_located((By.ID, "ddlRepairFunction")))
        selectlist3 = Select(select_OTHER)
        selectlist3.select_by_visible_text("[OTHER]")
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        enable_input2 = wait.until(EC.presence_of_element_located((By.NAME, "ttbDefectOperation")))
        last_column = enable_input2.get_attribute('value')
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        select = Select(select_list)
        select.select_by_visible_text("AC2")
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "ddlReturnOperation")))
        select = Select(select_list2)
        select.select_by_visible_text(last_column)
        input2 = wait.until(EC.presence_of_element_located((By.ID, "ttbRepairTime")))
        input2.send_keys("0")
        time.sleep(1)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()
        time.sleep(1)    
    finally:
        driver.quit()

def bug_c_automation(work_order,input_QR):
    global actions2
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument('--headless')    
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })  
    options.binary_location = chrome_path
    input_acc = "14575"
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=service)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="QryProg"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="在製品查詢"]')))
        actions.move_to_element(button).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        element6 = wait2.until(EC.presence_of_element_located((By.ID, "btnLotQuery")))
        element6.click()
        iframe2 = wait.until(EC.presence_of_element_located((By.XPATH, '//iframe[contains(@src, "LotListFieldSelect.aspx")]')))
        driver.switch_to.frame(iframe2)
        checkbox = wait2.until(EC.presence_of_element_located((By.ID, "tvn0CheckBox")))
        if checkbox.is_selected():
            checkbox.click()
        else:
            pass
        input_element3 = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl03_ttbValue")))
        input_element3.clear()
        input_element3.send_keys(input_QR)
        Query = wait.until(EC.presence_of_element_located((By.ID, "btnQuery")))
        Query.click()
        time.sleep(1)
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        if work_order in ["ARG","AMG"]:
            try:
                elementQR = wait.until(EC.presence_of_element_located((By.XPATH, f"//td[text()='{input_QR}']")))
                elementQR.click()
            except:
                error_code = "1"
                return error_code
        elif work_order == "AM2":
            try:
                element7 = wait.until(EC.presence_of_element_located((By.ID, "gvDataViewer_ctl02_hlLot")))
                driver.execute_script("arguments[0].click();", element7)
                actions.move_to_element(element7).click().perform()
                Niframes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
                driver.switch_to.frame(Niframes[0])
                element8 = wait.until(EC.presence_of_element_located((By.ID, "btnExit")))
                element8.click()
            except:
                error_code = "1"
                return error_code
        time.sleep(1)
        driver.switch_to.default_content()
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule_Modify"]')))
        element4.click()    
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '變更工作站')]")))
        actions.move_to_element(button2).click().perform()
        time.sleep(1)
        driver.switch_to.default_content()
        Giframes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
        driver.switch_to.frame(Giframes[1])
        select_list = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsOperation")))
        select = Select(select_list)
        select.select_by_index(2)
        time.sleep(0.5)
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsReason")))
        select2 = Select(select_list2)
        select2.select_by_index(1)
        time.sleep(0.5)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()
        time.sleep(5)    
    finally:
        driver.quit()

def bug_AR1_automation(input_QR):
    global workstation_name
    global actions2
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument('--headless')    
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })
    options.binary_location = chrome_path    
    input_acc = "14575"
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=service)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        driver.get("http://cimes.seec.com.tw/ARG/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="QryProg"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="在製品查詢"]')))
        actions.move_to_element(button).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        element5 = wait2.until(EC.element_to_be_clickable((By.XPATH, '//a[text()="批號"]')))
        element6 = driver.find_element(By.ID, "btnLotQuery")
        element6.click()
        iframe2 = wait.until(EC.presence_of_element_located((By.XPATH, '//iframe[contains(@src, "LotListFieldSelect.aspx")]')))
        driver.switch_to.frame(iframe2)
        checkbox = wait.until(EC.presence_of_element_located((By.ID, "tvn0CheckBox")))
        if checkbox.is_selected():
            checkbox.click()
        else:
            pass
        input_LOT = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl02_ttbValue")))
        input_LOT.clear()
        input_SN = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl03_ttbValue")))
        input_SN.clear()
        input_BOX = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl04_ttbValue")))
        input_BOX.clear()
        input_Machine = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl05_ttbValue")))
        input_Machine.clear()
        input_TYPE = wait.until(EC.presence_of_element_located((By.ID, "gvFilter_ctl08_ttbValue")))
        input_TYPE.clear()
        time.sleep(0.3)
        checkbox0 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_0")))
        if checkbox0.is_selected():
            checkbox0.click()
        else:
            pass
        checkbox1 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_1")))
        if checkbox1.is_selected():
            checkbox1.click()
        else:
            pass
        checkbox2 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_2")))
        if checkbox2.is_selected():
            checkbox2.click()
        else:
            pass
        checkbox3 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_3")))
        if checkbox3.is_selected():
            checkbox3.click()
        else:
            pass
        checkbox4 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_4")))
        if checkbox4.is_selected():
            checkbox4.click()
        else:
            pass
        checkbox5 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_5")))
        if checkbox5.is_selected():
            checkbox5.click()
        else:
            pass
        checkbox6 = wait.until(EC.presence_of_element_located((By.ID, "cblStatus_6")))
        if checkbox6.is_selected():
            checkbox6.click()
        else:
            pass
        input_SN.send_keys(input_QR)
        Query = wait.until(EC.presence_of_element_located((By.ID, "btnQuery")))
        Query.click()
        time.sleep(1)
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        elementQR = wait.until(EC.presence_of_element_located( (By.ID, "gvDataViewer_ctl02_hlLot")))
        elementQR.click()
        div_element = wait.until(EC.presence_of_element_located(
            (By.ID, "divOpenDialog")
        ))
        iframe_element = div_element.find_element(By.TAG_NAME, "iframe")
        driver.switch_to.frame(iframe_element)
        History = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "a[targer-panel-id='vHistory']")
        ))
        History.click()
        btnHisQuery = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "a#btnHisQuery.CimesButton")
        ))
        driver.execute_script("arguments[0].click();", btnHisQuery)
        page_number = 1
        last_button = None
        try:
            try:
                current_page = wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='current']")))
                print(f"找到本頁按鈕: {current_page.text}")
            except:
                print("找不到本頁按鈕")
                driver.quit()
            buttons = driver.find_elements(By.XPATH, "//a[@class='button'] | //span[@class='current']")
            if not buttons:
                print("找不到任何按鈕")
                driver.quit()
            page_numbers = [btn.text for btn in buttons]
            print(f"找到的頁碼: {page_numbers}")
            last_button = buttons[-1]
            print(f"最後一個按鈕為: {last_button.text}")
            last_button.click()
            print("成功點擊最後一個按鈕")
        except Exception as e:
            print(f"發生錯誤：{e}")
        time.sleep(2)
        rows = wait.until(EC.presence_of_all_elements_located(
            (By.XPATH, "//tr[@rowtype='DataRow']")
        ))
        last_row = rows[-1]
        workstation_name = last_row.find_elements(By.TAG_NAME, "td")[4].text
        print(f"最後一筆的工作站名稱為：{workstation_name}")
        btnExit = wait.until(EC.presence_of_element_located((By.ID, "btnExit")))
        btnExit.click()  
        driver.switch_to.default_content()
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule_Modify"]')))
        element4.click()    
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '變更工作站')]")))
        actions.move_to_element(button2).click().perform()
        time.sleep(1)
        driver.switch_to.default_content()
        Giframes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
        driver.switch_to.frame(Giframes[1])
        select_list = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsOperation")))
        select = Select(select_list)
        matched_option = None
        for option in select.options:
            option_text = option.text
            if workstation_name in option_text:
                matched_option = option_text
                break
        if matched_option:
            select.select_by_visible_text(matched_option)
            print(f"成功選擇工作站：{matched_option}")
        else:
            pass
        time.sleep(0.5)
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "_ddlcsReason")))
        select2 = Select(select_list2)
        select2.select_by_index(1)
        time.sleep(0.5)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()  
        time.sleep(2)
    finally:
        driver.quit()

def repair_automation(work_order,input_QR,class_number):
    global actions2
    global error_code
    global error_text
    error_text = None
    error_code = False
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument('--headless')    
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })     
    options.binary_location = chrome_path 
    input_acc = "14575"
    serice = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=serice)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        global last_column
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="WipRule"]')))
        element4.click()
        button2 = wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(@title, '維修(客)')]")))
        actions.move_to_element(button2).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        input1 = wait.until(EC.presence_of_element_located((By.ID, "ttbLot")))
        input1.send_keys(input_QR)
        input1.send_keys(Keys.RETURN)
        time.sleep(1)
        select_OTHER = wait.until(EC.presence_of_element_located((By.ID, "ddlRepairFunction")))
        selectlist3 = Select(select_OTHER)
        selectlist3.select_by_visible_text("[OTHER]")
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        enable_input2 = wait.until(EC.presence_of_element_located((By.NAME, "ttbDefectOperation")))
        last_column = enable_input2.get_attribute('value')
        select_list = wait.until(EC.presence_of_element_located((By.ID, "ddlDutyUnit")))
        select = Select(select_list)
        select.select_by_visible_text(class_number)
        select_list2 = wait.until(EC.presence_of_element_located((By.ID, "ddlReturnOperation")))
        select = Select(select_list2)
        select.select_by_visible_text(last_column)
        input2 = wait.until(EC.presence_of_element_located((By.ID, "ttbRepairTime")))
        input2.send_keys("0")
        time.sleep(2)
        finish_btnOK = wait.until(EC.presence_of_element_located((By.ID, "btnOK")))
        finish_btnOK.click()
        time.sleep(1)    
    finally:
        driver.quit()

def work_order_automation(work_order,data_input):
    global wait2,actions2
    global value
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument('--headless')    
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })      
    options.binary_location = chrome_path 
    input_acc = "14575"
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=service)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 20)  
    try:
        driver.get(f"http://cimes.seec.com.tw/{work_order}/Security/CimesUserLogin.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles

        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="CustRule"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="日工單開立"]')))
        actions.move_to_element(button).click().perform()
        time.sleep(3)
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        input_element3 = wait.until(EC.presence_of_element_located((By.ID, "CimesInputBox")))
        keyboard.press_and_release('ctrl+-')
        input_element3.send_keys(data_input)
        input_element3.send_keys(Keys.RETURN)
        time.sleep(2)
        enable_input = wait.until(EC.presence_of_element_located((By.ID, "ttbUnCreateQty")))
        value = enable_input.get_attribute('value')
        print(type(value))
        if value == '0':
            return value
        input_element4 = wait.until(EC.presence_of_element_located((By.ID, "ttbDayWOQty")))
        input_element4.send_keys(value)
        add_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btnAdd"))
        )                
        add_button.click()
        ok_button = wait.until(
        EC.element_to_be_clickable((By.ID, "btnOK"))
        )             
        actions.move_to_element(ok_button).click().perform()
        time.sleep(3)    
    finally:
        driver.quit()

def flowdata(work_order,input_QR):
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')  
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument('--headless') 
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })  
    options.binary_location = chrome_path
    input_acc = "14575"
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options,service=service)
    driver.set_window_size(1920, 1080)
    driver.maximize_window()
    actions2 = ActionChains(driver)
    wait = WebDriverWait(driver, 10)  
    wait2 = WebDriverWait(driver, 300)  
    try:
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        op_element = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navRULE.ctgr_hov"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(op_element).click().perform()
        element4 = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[righttype="WIPRULE"][cimesclass="QryProg"]')))
        element4.click()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="流程站點資訊查詢"]')))
        actions.move_to_element(button).click().perform()
        driver.switch_to.default_content()
        iframe = driver.find_element(By.CLASS_NAME, "win1")
        driver.switch_to.frame(iframe)
        checkbox1 = wait.until(EC.presence_of_element_located((By.ID, "gvQueryCondition_ctl02_cbxCheck")))
        if checkbox1.is_selected():
            checkbox1.click()
        else:
            pass
        checkbox2 = wait.until(EC.presence_of_element_located((By.ID, "gvQueryCondition_ctl03_cbxCheck")))
        if checkbox2.is_selected():
            checkbox2.click()
        else:
            pass
        checkbox3 = wait.until(EC.presence_of_element_located((By.ID, "gvQueryCondition_ctl04_cbxCheck")))
        if checkbox3.is_selected():
            checkbox3.click()
        else:
            pass
        checkbox3 = wait.until(EC.presence_of_element_located((By.ID, "gvQueryCondition_ctl04_cbxCheck")))
        if checkbox3.is_selected():
            checkbox3.click()
        else:
            pass
        checkbox4 = wait.until(EC.presence_of_element_located((By.ID, "gvQueryCondition_ctl06_cbxCheck")))
        if checkbox4.is_selected():
            checkbox4.click()
        else:
            pass
        input_element = wait.until(EC.presence_of_element_located((By.ID, "ttbDeviceFilter")))
        input_element.send_keys(input_QR)
        input_element.send_keys(Keys.RETURN)
        time.sleep(0.5)
        btnQuery = wait.until(EC.presence_of_element_located((By.ID, "btnQuery")))
        btnQuery.click()
        time.sleep(0.5)
        div_detail = driver.find_element(By.ID, "divDetail")
        station_blocks = div_detail.find_elements(By.CSS_SELECTOR, "ul > li")
        result = []
        for station in station_blocks:
            try:
                station_title = station.find_element(By.TAG_NAME, "h4").text.strip()
                tables = station.find_elements(By.CSS_SELECTOR, "table.CimesGridViewStyle")

                material_rows = []
                accessory_rows = []

                for i, table in enumerate(tables):
                    try:
                        rows = table.find_elements(By.CSS_SELECTOR, "tbody tr")
                        parsed_rows = []
                        for row in rows:
                            cells = row.find_elements(By.TAG_NAME, "td")
                            if i == 0 and len(cells) >= 2:
                                parsed_rows.append([cells[0].text.strip(), cells[1].text.strip()])
                            elif i == 1 and len(cells) >= 4:
                                parsed_rows.append([cells[0].text.strip(), cells[3].text.strip()])
                        if i == 0:
                            material_rows = parsed_rows
                        else:
                            accessory_rows = parsed_rows
                    except:
                        continue
                result.append({
                    "工作站": station_title,
                    "物料": material_rows,
                    "治具": accessory_rows
                })
            except Exception as e:
                print("錯誤略過：", e)
                continue
        from pprint import pprint
        pprint(result, width=160)
    finally:
        driver.quit()
        return result
    
def number_id_query(work_order,machine_id):
    chrome_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
    options = Options()
    options.add_argument("--disable-usb-devices")
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--ignore-certificate-errors')  
    options.add_argument('--disable-web-security')     
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")
    options.add_argument('--headless')
    options.add_argument("--disable-features=PasswordManagerEnabled,PasswordLeakDetection,AutofillServerCommunication")
    options.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    })  
    options.binary_location = chrome_path
    try:
        input_acc = "14575"
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(options=options,service=service)
        driver.set_window_size(1920, 1080)
        driver.maximize_window()
        actions2 = ActionChains(driver)
        wait = WebDriverWait(driver, 10)  
        wait2 = WebDriverWait(driver, 2)  
        driver.get(f"http://cimes.seec.com.tw/{work_order}/CimesDesktop.aspx")
        time.sleep(0.5)
        main_window = driver.current_window_handle
        for handle in driver.window_handles:
            if handle != main_window:
                new_window = handle
                break
        driver.switch_to.window(new_window)
        driver.close()
        driver.switch_to.window(main_window)
        username_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input#UserName"))
        )
        password_input = driver.find_element(By.CSS_SELECTOR, "input#Password")
        login_button = driver.find_element(By.ID, "LoginButton")
        username_input.send_keys(input_acc)
        password_input.send_keys(input_acc)
        login_button.click()
        
        all_windows = driver.window_handles
        for window in driver.window_handles:
            if window != all_windows:
                driver.switch_to.window(window)
                break
        element_menu = wait.until(
            EC.element_to_be_clickable((By.ID, "TestMenu"))
        )
        element_menu.click()
        driver.switch_to.frame("ifmMenu")
        element3 = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#navFRE.ctgr"))
        )
        actions = ActionChains(driver)
        actions.move_to_element(element3).click().perform()
        button = wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="人員上工查詢"]')))
        ActionChains(driver).move_to_element(button).click().perform()
        driver.switch_to.default_content()
        driver.switch_to.frame(0)
        filter_button = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR,
            "table#gvQuery thead tr th:first-child input[type='image']"
        )))
        filter_button.click()
        eqpid_input = wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR,
            "div[role='dialog'][aria-describedby='QueryTable'] input#gvQuery_query_EQPID"
        )))
        eqpid_input.clear()
        eqpid_input.send_keys(machine_id)
        btn = driver.find_element(By.XPATH, "//a[@class='CimesButton']//span[text()='查詢']")
        btn.click()
        time.sleep(1.5)
        table = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "table.CimesGridViewStyle")
        ))
        rows = table.find_elements(By.XPATH, ".//tr[@rowtype='DataRow']")
        result_lines = []
        for row in rows:
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) >= 4:
                eqpid = tds[1].text.strip()
                staff_id = tds[2].text.strip()
                staff_name = tds[3].text.strip()
                result_lines.append(f"{eqpid}｜{staff_id}｜{staff_name}")
        result_str = "\n".join(result_lines)
        return result_str
    finally:
        time.sleep(2)
        driver.quit      

def format_line_message(data):
    lines = []
    for item in data:
        section = f"工作站：{item['工作站']}\n"
        if item["物料"]:
            section += "物料："
            for m in item["物料"]:
                section += f"\n- {m[0]}（{m[1]}）"
        else:
            section += "物料：無"
        if item["治具"]:
            section += "\n治具："
            for j in item["治具"]:
                section += f"\n- {j[0]}（{j[1]}）"
        else:
            section += "\n治具：無"
        section += "\n"  
        lines.append(section)
    full_message = "\n".join(lines)
    if len(full_message) < 1000:
        return [full_message]
    else:
        return split_long_message(full_message)
     
def split_long_message(message, limit=950):
    parts = []
    while len(message) > limit:
        split_index = message.rfind("\n", 0, limit)
        if split_index == -1:
            split_index = limit
        parts.append(message[:split_index].strip())
        message = message[split_index:].strip()
    if message:
        parts.append(message)
    return parts

def image(user_id,content):
    image_url = f"https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={content}"
    image_message = ImageSendMessage(
    original_content_url=image_url,  
    preview_image_url=image_url     
)
    line_bot_api.push_message(user_id, image_message)

def select_web(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='部門選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/Qj1b4KL/2.jpg',
                        title='部門選擇',
                        text='請從以下選項選擇一個部門：',
                        actions=[
                            MessageAction(label='ARG', text='ARG'),
                            MessageAction(label='AMG', text='AMG'),
                            MessageAction(label='AM2', text='AM2')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇部門，或輸入[返回]結束動作")
            ]
        )
    
def select_web2(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='班別選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/cyBPJHW/2.jpg',
                        title='班別選擇',
                        text='請從以下選項選擇一個班別：',
                        actions=[
                            MessageAction(label='AF1', text='AF1'),
                            MessageAction(label='AF2', text='AF2'),
                            MessageAction(label='AF3', text='AF3')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
            ]
        )

def select_web3(tk):
    line_bot_api.reply_message(
            tk,
            [
                TemplateSendMessage(
                    alt_text='班別選擇',
                    template=ButtonsTemplate(
                        thumbnail_image_url='https://i.ibb.co/KFFgHK1/3.jpg',
                        title='班別選擇',
                        text='請從以下選項選擇一個班別：',
                        actions=[
                            MessageAction(label='AC1', text='AC1'),
                            MessageAction(label='AC2', text='AC2'),
                            MessageAction(label='AC3', text='AC3')
                        ]
                    )
                ),
                TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
            ]
        )

def select_web4(tk):
    line_bot_api.reply_message(
        tk,
        [
            TemplateSendMessage(
                alt_text='班別選擇 1',
                template=ButtonsTemplate(
                    thumbnail_image_url='https://i.ibb.co/z2sW8TF/image.jpg',
                    title='班別選擇 1',
                    text='請從以下選項選擇一個班別：',
                    actions=[
                        MessageAction(label='AR1', text='AR1'),
                        MessageAction(label='AR2', text='AR2'),
                        MessageAction(label='AR3', text='AR3')
                    ]
                )
            ),
            TemplateSendMessage(
                alt_text='班別選擇 2',
                template=ButtonsTemplate(
                    thumbnail_image_url='https://i.ibb.co/z2sW8TF/image.jpg',  
                    title='班別選擇 2',
                    text='請從以下選項選擇其他班別：',
                    actions=[
                        MessageAction(label='AR4', text='AR4'),
                        MessageAction(label='AF1', text='AF1'),
                        MessageAction(label='返回', text='返回')
                    ]
                )
            ),
            TextSendMessage(text="請選擇班別，或輸入[返回]結束動作")
        ]
    )

def send_carousel_message(user_id):
    line_bot_api.push_message(user_id, TemplateSendMessage(
        alt_text='Menu',
        template=CarouselTemplate(
            columns=[
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/Qj1b4KL/2.jpg',
                    title='常用功能',
                    text='以下為常用功能',
                    actions=[
                        MessageAction(
                            label='工單查詢',
                            text='工單查詢'
                        ),
                        MessageAction(
                            label='AC2專用',
                            text='AC2專用'
                        ),
                        MessageAction(
                            label='飛輪專用',
                            text='飛輪專用'
                        )
                    ]
                ),
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/Ngq4L8h/Manufacturing-Execution-System-2.jpg',  
                    title='查詢服務',
                    text='以下為查詢相關服務',
                    actions=[
                        MessageAction(
                            label='人員上工',
                            text='人員上工'
                        ),
                        MessageAction(
                            label='流程查詢',
                            text='流程查詢'
                        ),
                        MessageAction(
                            label='QRcode產生器',
                            text='QRcode小工具'
                        )
                    ]
                ),
                CarouselColumn(
                    thumbnail_image_url='https://i.ibb.co/mHd3tpR/2.jpg',  
                    title='其他功能',
                    text='以下為其他功能',
                    actions=[
                        MessageAction(
                            label='開立工單',
                            text='開立工單'
                        ),
                        MessageAction(
                            label='物料QR產生器',
                            text='物料QR小工具'
                        ),
                        MessageAction(
                            label='解除維修站',
                            text='解除維修站'
                        ),
                    ]
                )
            ]
        )
    ))

def replace_number_with_random_iteratively(user_id,text,tk):
    parts = text.split(";")
    if len(parts) < 7 or not parts[5].isdigit():
        line_bot_api.reply_message(tk, TextSendMessage("格式錯誤"))
        return text, None
    original_box_number = parts[5]
    while True:
        generated_numbers = str(random.randint(10000, 99999))
        if generated_numbers != original_box_number:
            break
    parts[5] = generated_numbers
    edited_text = ";".join(parts)
    image(user_id=user_id, content=edited_text)
    return edited_text, generated_numbers


def log_action(id="", event="",department="", class_number="",Serial_number="", Result_output=""):
    check_or_create_excel(file_path)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp, id, event, department, class_number, Serial_number, Result_output])
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    workbook.save(file_path)
    print(f"檔案 '{file_path}' 已更新並儲存。")

def find_work_order_by_part(part_number):
    current_dir2 = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir2, "工單.xlsx") 
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}
        required_fields = ["工單號碼", "件號", "摘要"]
        if not all(field in headers for field in required_fields):
            print("Excel 欄位需包含：工單號碼、件號、摘要")
            return None, None
        work_order_col = headers["工單號碼"]
        part_number_col = headers["件號"]
        summary_col = headers["摘要"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[part_number_col]) == str(part_number):
                work_order = row[work_order_col]
                summary = row[summary_col]
                return work_order, summary
        print("找不到工單。")
        return None, None
    except FileNotFoundError:
        print(f"找不到檔案：{file_path}")
        return None, None
    
current_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_dir, "line_bot_log.xlsx")  
check_or_create_excel(file_path)

app = Flask(__name__)

@app.route("/", methods=['POST'])
def linebot():
    global class_process
    global class_number
    global work_order
    global user_ticket_number
    global in_process
    global in_process2
    global pass_process
    global sec_in_process
    global sec_in_process2
    global user_ticket_number2
    global unlock_repaire_station
    global unlock_repaire_station_memory
    global unlock_repaire_station2
    global QRcode_check
    global QRcode_content
    global return_msg
    global internet_process
    global internet_process2
    global Material
    global title_value
    global error_code
    global value
    global image_process
    global search_value
    global paper
    global internet_process3
    global quantity
    global event
    global share_url
    global flow1
    global flow2
    global work_flow1
    global work_flow2
    global user_state
    global work_check1
    global AC2_flow
    body = request.get_data(as_text=True)                    

    try:
        json_data = json.loads(body)                         
        access_token = 'Pn6oyBEK0RJLR1U0OEY7q35O+RTM0PaHrSAzQ8805ZRqNqf/O2A1CM5YWMeQoB+8ngYCLqoEK6nmHi9T6YCycuSOSKHiTBQWG56vUUY6RTbDr96z2Iq+Uo7SSVdWWBuDOdClGrB4Y8bfvXrMqmnwBQdB04t89/1O/w1cDnyilFU='
        secret = '7599d376613784d49a0230e1b53a7eaf'
        line_bot_api = LineBotApi(access_token)              
        handler = WebhookHandler(secret)                    
        signature = request.headers['X-Line-Signature']      
        handler.handle(body, signature)                      
        tk = json_data['events'][0]['replyToken']          
        type = json_data['events'][0]['message']['type']  
        user_id = None   
        user_id = json_data['events'][0]['source']['userId']
        
        if type == "text" :
            msg = json_data['events'][0]['message']['text']  
            if msg == "選單":
                send_carousel_message(user_id)    
            elif msg == "開立工單":
                event = msg
                select_web(tk)
                value = None
                title_value = None
                error_code = None
                user_ticket_number = None
                work_order = None
                in_process = True 
            elif in_process:
                if msg == "返回":
                    in_process = False
                else:
                    line_bot_api.push_message(user_id, TextSendMessage("請輸入工單號碼，或輸入[返回]結束動作"))                                            
                    work_order = msg        
                    sec_in_process = True
                    in_process = False 
            elif sec_in_process:
                if msg == "返回":
                    sec_in_process = False
                else:                       
                    user_ticket_number = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"工單號碼為：{user_ticket_number}，處理中，請稍候"))                
                    try:
                        work_order_automation(work_order = work_order,data_input = user_ticket_number)
                        if value == '0':
                            image(user_id=user_id,content=f"{user_ticket_number}-001")
                            line_bot_api.push_message(user_id, TextSendMessage("工單在月初已開立"))
                        else:    
                            image(user_id=user_id,content=f"{user_ticket_number}-001")
                            line_bot_api.push_message(user_id, TextSendMessage("🎉🎉🎊工單開立完成🎊🎉🎉"))
                            log_action(id=user_id, event="開立工單",department=work_order,Serial_number=value)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("查無此工單"))
                        pass
                    sec_in_process = False
            elif msg == "飛輪專用":
                event = msg
                user_ticket_number2 = None
                line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                sec_in_process2 = True
            elif sec_in_process2:
                if msg == "返回":
                    sec_in_process2 = False
                else:
                    user_ticket_number2 = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"序號：{user_ticket_number2}，處理中，請稍候"))
                    try:
                        bug_AR1_automation(input_QR=user_ticket_number2)
                        line_bot_api.push_message(user_id, TextSendMessage(f"🎉🎊解除維修完成🎊🎉，請從「{workstation_name}」刷讀")) 
                        log_action(id=user_id, event="飛輪專用",department="AR1",Serial_number=user_ticket_number2)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("系統發生異常")) 
                    sec_in_process2 = False       
            elif msg == "解除維修站":
                event = msg
                select_web(tk)
                unlock_repaire_station = True
            elif unlock_repaire_station:
                if msg == "返回":
                    unlock_repaire_station = False
                elif msg == "ARG":
                    work_order = msg
                    select_web4(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                elif msg == "AMG":
                    work_order = msg
                    select_web3(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                elif msg == "AM2":
                    work_order = msg
                    select_web2(tk)
                    unlock_repaire_station = False
                    unlock_repaire_station2 = True
                else:    
                    select_web(tk)
            elif unlock_repaire_station2:
                if msg == "返回":
                    unlock_repaire_station2 = False
                elif work_order == "ARG" and (msg == "AR1" or msg == "AR2" or msg == "AR3" or msg == "AR4" or msg == "AF1"):
                    class_number = msg                  
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_process = True
                elif work_order == "AMG" and (msg == "AC1" or msg == "AC2" or msg == "AC3"):
                    class_number = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_process = True
                elif work_order == "AM2" and (msg == "AF1" or msg == "AF2" or msg == "AF3"):
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                    unlock_repaire_station2 = False
                    class_number = msg
                    class_process = True
                else:
                    if work_order == "ARG":                      
                        select_web4(tk)
                    elif work_order == "AMG":
                        select_web3(tk)
                    elif work_order == "AM2":
                        select_web2(tk)
                    else:
                        pass
            elif class_process:
                if msg == "返回":
                    class_process = False
                else:
                    unlock_repaire_station_memory = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"{unlock_repaire_station_memory}處理中，請稍後"))
                    try:
                        error_code = None
                        repair_automation(work_order = work_order,input_QR = unlock_repaire_station_memory,class_number=class_number)
                        line_bot_api.push_message(user_id, TextSendMessage(f"🎉🎉🎊解除維修完成🎊🎉🎉，請從「{last_column}」刷讀")) 
                        log_action(id=user_id, event="解除維修站",department=work_order, class_number=class_number,Serial_number=unlock_repaire_station_memory)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("系統發生異常")) 
                        pass
                    class_process = False
            elif error_code:
                line_bot_api.push_message(user_id, TextSendMessage(F"{error_text}"))
                error_code = False
            elif msg == "QRcode小工具":     
                event = msg
                line_bot_api.reply_message(tk, TextSendMessage("請輸入內容，或輸入[返回]取消動作"))
                QRcode_check = True
            elif QRcode_check:
                if msg == "返回":
                    QRcode_check = False
                else:
                    QRcode_content = msg
                    image(user_id=user_id,content=QRcode_content)
                    QRcode_check = False
                    try:
                        log_action(id=user_id, event="QRcode小工具", Serial_number=QRcode_content)
                        print("成功")
                    except:
                        print("失敗")
            elif msg == "物料QR小工具":
                event = msg
                line_bot_api.reply_message(tk, TextSendMessage("請輸入材料編號，或輸入[返回]結束動作"))
                Material = True
            elif Material:
                if msg == "返回":
                    Material = False
                else:
                    text = msg
                    edited_text, generated_numbers = replace_number_with_random_iteratively(user_id=user_id,text=text,tk=tk)
                    log_action(id=user_id, event="物料QR小工具", Serial_number=text, Result_output=edited_text)
                    Material = False  
            elif msg == "流程查詢":
                event = msg
                select_web(tk)
                flow1 = True
            elif flow1:
                if msg == "返回":
                    flow1 = False
                else:
                    work_order = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入件號，或輸入[返回]結束動作"))
                    flow1 = False
                    flow2 = True
            elif flow2:
                if msg == "返回":
                    flow2 = False
                else:
                    try:
                        input_QR = msg
                        line_bot_api.reply_message(tk, TextSendMessage("請稍後~")) 
                        result = flowdata(work_order,input_QR)
                        result2 = format_line_message(data=result)
                        print(result2)
                        if isinstance(result2, list):
                            result2 = '\n'.join(result2)
                        line_bot_api.push_message(user_id, TextSendMessage(text=result2))
                        log_action(id=user_id, event="流程查詢",department=work_order,Serial_number=input_QR)
                        flow2 = False
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage(F"找不到{input_QR}的流程")) 
                        flow2 = False
            elif msg == "人員上工":
                event = msg
                select_web(tk)
                work_flow1 = True
            elif work_flow1:
                if msg == "返回":
                    work_flow1 = False
                else:
                    work_order = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請輸入設備編號，或輸入[返回]結束動作"))                
                    work_flow2 = True
                    work_flow1 = False
            elif work_flow2:
                if msg == "返回":
                    work_flow2 = False
                else:
                    machine_id = msg
                    line_bot_api.reply_message(tk, TextSendMessage("請稍後~"))  
                    result_str = number_id_query(work_order,machine_id)
                    line_bot_api.push_message(user_id, TextSendMessage(result_str)) 
                    log_action(id=user_id, event="人員上工",department=work_order,Serial_number=result_str)
                    work_flow2 = False
            elif msg == "工單查詢":
                event = msg
                line_bot_api.reply_message(tk, TextSendMessage("請輸入件號，或輸入[返回]結束動作"))
                work_check1 = True
            elif work_check1:
                if msg == "返回":
                    work_check1 = False
                else:
                    input_part = msg
                    try:
                        wo, summary = find_work_order_by_part(input_part)
                        if wo is not None:
                            line_bot_api.reply_message(tk, TextSendMessage(f"工單號碼：{wo}，摘要：{summary}"))  
                            log_action(id=user_id, event="工單查詢", Serial_number=input_part, Result_output=wo)
                            text = f"{wo}-001"
                            image(user_id=user_id,content=text)
                            work_check1 = False
                        else:
                            line_bot_api.reply_message(tk, TextSendMessage(f"找不到工單。"))  
                            work_check1 = False
                    except:
                        line_bot_api.reply_message(tk, TextSendMessage(f"系統異常"))  
                        work_check1 = False
            elif msg == "AC2專用":
                event = msg
                SN_input = None
                line_bot_api.reply_message(tk, TextSendMessage("請輸入序號，或輸入[返回]結束動作"))
                AC2_flow = True
            elif AC2_flow:
                if msg == "返回":
                    AC2_flow = False
                else:
                    SN_input = msg
                    line_bot_api.reply_message(tk, TextSendMessage(f"序號：{SN_input}，處理中，請稍候"))
                    try:
                        repair_ac2_automation(SN_input=SN_input)
                        line_bot_api.push_message(user_id, TextSendMessage(f"🎉🎊解除維修完成🎊🎉，請從「{workstation_name}」刷讀")) 
                        log_action(id=user_id, event="AC2專用",department="AC2",Serial_number=SN_input)
                    except:
                        line_bot_api.push_message(user_id, TextSendMessage("系統發生異常")) 
                    AC2_flow = False  
            else:
                line_bot_api.reply_message(tk,TextSendMessage('其他需求請洽：製管MES組'))
        else:
            line_bot_api.reply_message(tk,TextSendMessage('你傳的不是文字呦～'))
    except:
        print(body)                                         
    return 'OK'                                              

if __name__ == "__main__":
    app.run()
